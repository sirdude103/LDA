from gensim import corpora, models # dependency for use of bag of words
import json # dependency to work with JSON input
import openpyxl # dependency to read/write to Excel document
import shutil # dependency used to copy Excel documents


# python processing files
from tokenization import tokenize
from stopping import remove_stops
from stemming import stem_words

# make sure to pip install the following:
## gensim
## nltk
## stop_words

###################### FUNCTIONS #########################

# Finds a column given its header
# header_row : xlrd.sheet.row : the row where the column headers are located
# column_name : string : the value of the column header located in row 0
#
# returns : int : the int of the column it's located in
def find_column_by_value(header, column_value):

    body_column = 0
    for column_title in header:
        if column_title.value == column_value:
            break
        body_column += 1
    else: 
        raise Exception('"{}" was not found in the header given.'.format(column_value))
    
    return body_column

    
# Retrieves documents from an Excel document column
# file_name : string : given location of an excel file and column
#
# returns : xlrd.sheet : the first xlrd worksheet in the given file
def retrieve_sheet_from_excel(file_name):
    workbook = openpyxl.load_workbook(filename=file_name)
    return workbook.worksheets[0]
    
    

# Retrieves JSON from its file and returns as a set
# file_name : string : given location of .json file
#
# returns : dictionary : the json values
def retrieve_json(file_name):
    with open(file_name, "r") as json_file:
        return json.load(json_file)

        
# Extracts topics from the doc set
# doc_set : list of strings : the document bodies where topics should be read
# num_of_topics : int : the amount of different topics expected out of all the documents
# num_of_words : int : the amount of words returned to represent a topics
# 
# returns : list of tuples : the major topics and and representing words as a list of tuples
def run_extractor(doc_set, num_of_topics, num_of_words):
    raw = []
    
    if type(doc_set) == str:
        doc_set = [doc_set]
        
    for doc in doc_set:
    
        # odd scenario where our document text is given in a single element list
        if type(doc) == list:
            doc = doc[0]
        raw_doc = doc.lower()
        
        # convert to tokens
        tokens = tokenize( raw_doc )
        
        # remove stop words from tokens
        stopped_tokens = remove_stops( tokens )
        
        # stem token
        stemmed_tokens = stem_words( stopped_tokens )
        
        # add to our list
        raw.append( stemmed_tokens )
        
    dictionary = corpora.Dictionary( raw )
    
    # convert to bag of words 
    # tuples where tuple[0] is word key and tuple[1] is word occurence in given document
    corpus = [dictionary.doc2bow(text) for text in raw]
    
    ldamodel = models.ldamodel.LdaModel(corpus, num_topics=num_of_topics, id2word = dictionary, passes=20)
    
    return ldamodel.print_topics(num_topics=num_of_topics, num_words=num_of_words)


# Splits a single document into smaller blocks in order to be extracted
# doc : string : text content of a document
# blocks : int : desired amount of splitted blocks
# 
# returns : dictionary : the blocks of the document
def split_doc(doc, blocks):
    raise Exception("Function: 'split_doc' is currently not implemented.")
    
    
def topic_match(topic1, topic2):
    match_count = 0
    for word1 in topic1[1]:
        for word2 in topic2[1]:
            if word1 == word2:
                match_count += 1
    return match_count
    
    
# Takes a tuple in ldamodel bag of words format and unpacks it
# given_tuple : tuple : tuple to unpack in form of (int, 'float*"string"' + ...)
# 
# returns : list of 2 lists : first list is frequency of words. second list is the words
def unpack_tuple(given_tuple):
    freqs = []
    words = []
    output = [freqs, words]
    
    # We ignore the ID of the given tuple
    freq_word_combos = given_tuple[1].split(" + ")
    
    for combo in freq_word_combos:
        combo = combo.split("*")
        freqs.append(combo[0])
        words.append(combo[1])
    
    return output
    
###################### END FUNCTIONS ######################

###################### DEPRECATED CODE ####################

# # plaintext test documents
from test_docs import test_documents
from full_test_doc import full_doc

# retrieve text from JSON
#test_json = retrieve_json("FormattedJSONTest.json")
#document_texts = [doc["text"] for doc in test_json["documents"]]

#print("Type of document_texts: {}".format(type(document_texts)))
#topics = run_extractor(document_texts, 5, 3)

#for topic in topics:
#    print(topic)

#print("Type of full_doc: {}".format(type(full_doc)))
#print("Content: {}".format(full_doc))
#topics = run_extractor(full_doc, 5, 3)

#for topic in topics:
#    print(topic)   

#################### END DEPRECATED CODE ##################


def main():
    # Declare constants
    INPUT_FILE = "Column_Setup.xlsx"
    OUTPUT_FILE = "Topics.xlsx"
    HEADER_ROW = 0
    SINGLE_DOC_TOPICS = 5
    TOTAL_TOPICS = 5
    TOTAL_WORDS = 3
    VERBOSE = True
    
    # Retrieve documents from the excel file
    shutil.copy(INPUT_FILE, OUTPUT_FILE)
    
    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    worksheet = workbook.worksheets[0]
    
    column_titles = worksheet[HEADER_ROW+1]
    body_column_id = find_column_by_value(column_titles, "Body")
    document_texts = []
    for row_id in range(len(list(worksheet.rows))):
        if row_id != HEADER_ROW: 
            document_texts.append(worksheet.cell(row=row_id+1, column=body_column_id+1).value)
    
    documents_topics = []
    
    # Find the topics of each document
    current_row_id = HEADER_ROW
    for text in document_texts: 
        if VERBOSE: 
            print("Document texts: {}".format(text))
            print("Type of the variable: {}".format(type(text)))
            print("Topics: ")
        
        current_row_id += 1
        topics = run_extractor(text, SINGLE_DOC_TOPICS, TOTAL_WORDS)
        document_topics = []
        
        # unpack each topic
        for topic in topics:
            if VERBOSE: print(topic)
            document_topics.append(unpack_tuple(topic))
        
        documents_topics.append(document_topics)
        
        # add document_topics to excel document
        if VERBOSE:
            print("document_topics: ")
            print(document_topics)
            
        current_column_id = body_column_id
        for topic in document_topics:
            zippedTopic = list(zip(topic[0], topic[1]))
            
            if VERBOSE:
                print("topic: ")
                print(topic)
                print("zippedTopic: ")
                print(zippedTopic)
            
            for combo in zippedTopic:
                current_column_id += 1
                worksheet.cell(current_row_id+1, current_column_id+1).value = combo[0]
            
                current_column_id += 1
                worksheet.cell(current_row_id+1, current_column_id+1).value = combo[1]
            

    # Find the topics of all the documents together
    if VERBOSE: print("Topics of individual documents found. Now processing all together...")
    
    if VERBOSE: 
        print("Type of the variable: {}".format(type(document_texts)))
        print("Topics: ")
    
    topics = run_extractor(document_texts, TOTAL_TOPICS, TOTAL_WORDS)
    
    overarching_topics = []
    for topic in topics:
        if VERBOSE: print(topic)
        overarching_topics.append(unpack_tuple(topic))
    
    # Match document with its most accurate topic
    if VERBOSE: print(overarching_topics)
    
    # current nested for loop is processing len(documents_topics) * SINGLE_DOC_TOPICS times.
    documents_best_topic_id = []
    documents_most_related_topic_id = []
    
    for document_topics in documents_topics:
        doc_topic_id = -1
        
        doc_best_topic = 0 # ID of the document's best topic
        overaching_best_topic = 0 # ID of the connected overarching topic
        doc_best_match = 0 # Best matches out of every doc_topic-overarching_topic combo
    
        for document_topic in document_topics: # loops SINGLE_DOC_TOPICS times
            doc_topic_id += 1
            
            overarching_topic_id = -1
            overarching_best_id = 0
            overarching_best_match = 0
            
            for overarching_topic in overarching_topics:
                overarching_topic_id += 1
                current_match = topic_match(overarching_topic, document_topic)
                if current_match > overarching_best_match:
                    overarching_best_match = current_match
                    overarching_best_id = overarching_topic_id
                    if overarching_best_match == TOTAL_WORDS: 
                        break
                        
            if overarching_best_match > doc_best_match:
                doc_best_topic = doc_topic_id
                overarching_best_topic = overarching_best_id
                doc_best_match = overarching_best_match
                
                if doc_best_match == TOTAL_WORDS:
                    break
        
        documents_best_topic_id.append(doc_best_topic)
        documents_most_related_topic_id.append(overarching_best_topic)   
    
    # Display best matched topic
    if VERBOSE:    
        print("BEST MATCHING TOPICS: ")
        for i in range(len(documents_most_related_topic_id)):
            print(documents_most_related_topic_id[i])
    
    column_id = find_column_by_value(column_titles, "Best_Topic")
    
    for row_id in range(len(list(worksheet.rows))):
        if row_id != HEADER_ROW: 
            worksheet.cell(row=row_id+1, column=column_id+1).value = documents_best_topic_id[row_id-1] + 1
    
    column_id = find_column_by_value(column_titles, "Matching_Overarching")
    
    for row_id in range(len(list(worksheet.rows))):
        if row_id != HEADER_ROW: 
            worksheet.cell(row=row_id+1, column=column_id+1).value = documents_most_related_topic_id[row_id-1] + 1
    
    # Create new sheet to save overarching topics
    bt_sheet = workbook.create_sheet("Best_Topics", 1)
    
    for letter in "abcdefghijklmnopqrstuvwxyz".upper():
        bt_sheet.column_dimensions[letter].width = 20 
    
    for column_id in range(2, (TOTAL_WORDS*2)+2):
        if column_id % 2 == 0: 
            bt_sheet.cell(row=1, column= column_id).value = "Word_{}_Freq".format(int(column_id/2))
        else:
            bt_sheet.cell(row=1, column= column_id).value = "Word_{}_Word".format(int(column_id/2))
    row_id=0
    for overarching_topic in overarching_topics:
        row_id += 1
        bt_sheet.cell(row=row_id+1, column=1).value = "Topic {}:".format(row_id)
        zipped_topic = list(zip(overarching_topic[0], overarching_topic[1]))
        
        current_column_id = 0
        for combo in zipped_topic:
            current_column_id += 1
            bt_sheet.cell(row_id+1, current_column_id+1).value = combo[0]
        
            current_column_id += 1
            bt_sheet.cell(row_id+1, current_column_id+1).value = combo[1]
            
    
    # Save our changes
    if VERBOSE: print("Saving our changes...")
    workbook.save(OUTPUT_FILE)
    if VERBOSE: print("Changes saved to {}.".format(OUTPUT_FILE))
    
main()
